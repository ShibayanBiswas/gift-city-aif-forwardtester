"""Monte Carlo Nifty matrix — Excel ``Nifty Simulations.xlsx`` layout.

Rows = path numbers 1…N (vertical).
Columns = trading **dates** from as-of through Product End (tenure calendar end).

All Path / Start / End parameter rows share the same window (Start = as-of,
End = path_end_calendar). Same calendar date ⇒ different prices across paths
(independent Z per path_id). Each cell follows::

    S_t = S_{t-1} · exp(drift + σ · Z),  Z ~ N(0,1)

Path tenure windows slice this matrix so hedge / NAV / roll points stay
aligned to the same path_id random stream.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import numpy as np

from .gbm import GBM_BASE_SEED, GbmParams, gbm_spots, gbm_spots_matrix

_MONTHS = (
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)

# Soft cap note only — export always streams path-by-path with branded chrome.
_EXCEL_CELL_SOFT_CAP = 800_000

# Desk brand tokens — Primary SP Dashboard `export-theme.ts` (openpyxl RGB, no FF prefix).
_BRAND_MAROON = "7A1E2C"
_BRAND_MAROON_DEEP = "5C1622"
_BRAND_GOLD = "D4B24C"
_BRAND_GOLD_PALE = "FCF8EE"
_BRAND_IVORY = "FAF7EF"
_BRAND_PARCHMENT = "F8F4EA"
_BRAND_INK = "1C1917"
_BRAND_MUTED = "78716C"
_BRAND_WHITE = "FFFFFF"
_BRAND_BORDER = "E7E1CF"
_BRAND_RULE = "C9B88A"
_BRAND_FOOTER = "F8F4EA"

# Chrome layout (matches frontend/lib/download.ts + Primary SP masthead).
_CHROME_HEADER_ROW = 9  # Path / Start / End / dates
_CHROME_DATA_START = 10
_DESK_EYEBROW = "Anand Rathi Wealth · Gift City AIF Forwardtester"


def _resolve_logo() -> Path | None:
    """Logo for branded Excel — works in full repo and Render (backend-only root)."""
    here = Path(__file__).resolve()
    candidates = [
        here.parents[1] / "static" / "brand" / "arwl-logo.png",  # backend/app/static (Render)
        here.parents[3] / "frontend" / "public" / "brand" / "arwl-logo.png",  # monorepo
        here.parents[2] / "static" / "brand" / "arwl-logo.png",
    ]
    for p in candidates:
        if p.is_file():
            return p
    return None


_LOGO = _resolve_logo()


def _desk_date(d: date | str) -> str:
    """Match frontend formatDeskDate: 31-Jul-2026."""
    if isinstance(d, str):
        parts = d.strip()[:10].split("-")
        if len(parts) != 3:
            return d
        y, m, day = (int(parts[0]), int(parts[1]), int(parts[2]))
    else:
        y, m, day = d.year, d.month, d.day
    return f"{day:02d}-{_MONTHS[m - 1]}-{y}"


def horizon_trading_dates(market_dates: list[date], asof: date, horizon: date) -> list[date]:
    """Inclusive trading sessions from as-of through Simulation End."""
    return [d for d in market_dates if asof <= d <= horizon]


def build_mc_matrix(
    params: GbmParams,
    dates: list[date],
    n_paths: int,
    *,
    base_seed: int = GBM_BASE_SEED,
) -> np.ndarray:
    """Return float32 matrix shape (n_paths, n_dates); row i = path_id i+1."""
    if not dates or n_paths <= 0:
        return np.zeros((0, 0), dtype=np.float32)
    mat = gbm_spots_matrix(
        params.spot0,
        len(dates),
        n_paths,
        params.drift,
        params.std_dev,
        base_seed=base_seed,
    )
    return np.asarray(mat, dtype=np.float32)


def slice_path_spots(
    mat: np.ndarray,
    horizon_dates: list[date],
    path_dates: list[date],
    path_id: int,
) -> np.ndarray:
    """Slice one path's tenure spots from the full horizon matrix."""
    if mat.size == 0 or not path_dates:
        return np.zeros(0, dtype=float)
    idx = {d: i for i, d in enumerate(horizon_dates)}
    row = int(path_id) - 1
    if row < 0 or row >= mat.shape[0]:
        raise IndexError(f"path_id {path_id} out of matrix rows {mat.shape[0]}")
    return np.asarray([float(mat[row, idx[d]]) for d in path_dates if d in idx], dtype=float)


def spots_aligned_to_horizon(
    path_dates: list[date],
    params: GbmParams,
    path_id: int,
    horizon_dates: list[date],
    *,
    base_seed: int = GBM_BASE_SEED,
) -> np.ndarray:
    """Generate full horizon GBM row then slice to path dates (Excel column alignment)."""
    if not path_dates:
        return np.zeros(0, dtype=float)
    if not horizon_dates:
        return gbm_spots(
            params.spot0,
            len(path_dates),
            params.drift,
            params.std_dev,
            path_id=path_id,
            base_seed=base_seed,
        )
    full = gbm_spots(
        params.spot0,
        len(horizon_dates),
        params.drift,
        params.std_dev,
        path_id=path_id,
        base_seed=base_seed,
    )
    idx = {d: i for i, d in enumerate(horizon_dates)}
    missing = [d for d in path_dates if d not in idx]
    if missing:
        raise ValueError(
            f"Path tenure has {len(missing)} date(s) outside the GBM horizon "
            f"(first missing {missing[0].isoformat()})."
        )
    return np.asarray([float(full[idx[d]]) for d in path_dates], dtype=float)


def save_mc_matrix(
    folder: Path,
    *,
    dates: list[date],
    matrix: np.ndarray,
    params: GbmParams,
    base_seed: int,
) -> Path:
    """Persist matrix + dates under the job folder."""
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / "mc_matrix.npz"
    np.savez_compressed(
        path,
        matrix=np.asarray(matrix, dtype=np.float32),
        dates=np.array([d.isoformat() for d in dates]),
        spot0=float(params.spot0),
        drift=float(params.drift),
        std_dev=float(params.std_dev),
        mean_return=float(params.mean_return),
        base_seed=int(base_seed),
        asof=str(params.asof),
        first_date=str(params.first_date),
        last_date=str(params.last_date),
    )
    return path


def load_mc_matrix(folder: Path, *, mmap: bool = False) -> dict[str, Any] | None:
    path = folder / "mc_matrix.npz"
    if not path.exists():
        return None
    # mmap keeps huge grids off the heap when only previewing / streaming rows.
    data = np.load(path, allow_pickle=False, mmap_mode="r" if mmap else None)
    dates = [date.fromisoformat(str(x)) for x in np.asarray(data["dates"]).tolist()]
    keys = set(data.files)
    matrix = data["matrix"]
    if not mmap:
        matrix = np.asarray(matrix, dtype=np.float32)
    return {
        "matrix": matrix,
        "dates": dates,
        "spot0": float(data["spot0"]),
        "drift": float(data["drift"]),
        "std_dev": float(data["std_dev"]),
        "mean_return": float(data["mean_return"]),
        "base_seed": int(data["base_seed"]),
        "asof": str(data["asof"]),
        "first_date": str(data["first_date"]) if "first_date" in keys else "2001-01-01",
        "last_date": str(data["last_date"]) if "last_date" in keys else str(data["asof"]),
        "n_paths": int(matrix.shape[0]),
        "n_dates": int(matrix.shape[1]),
    }


def matrix_meta(payload: dict[str, Any]) -> dict[str, Any]:
    dates: list[date] = payload["dates"]
    mat = payload.get("matrix")
    n_paths = int(payload["n_paths"]) if "n_paths" in payload else int(mat.shape[0]) if mat is not None else 0
    n_dates = int(payload["n_dates"]) if "n_dates" in payload else (
        int(mat.shape[1]) if mat is not None else len(dates)
    )
    return {
        "asof": payload["asof"],
        "n_paths": n_paths,
        "n_dates": n_dates,
        "first_date": dates[0].isoformat() if dates else None,
        "last_date": dates[-1].isoformat() if dates else None,
        "spot0": payload["spot0"],
        "drift": payload["drift"],
        "std_dev": payload["std_dev"],
        "mean_return": payload["mean_return"],
        "base_seed": payload["base_seed"],
        "layout": {
            "rows": "path_id 1…N vertical",
            "columns": "trading dates as-of → Simulation End horizontal",
            "formula": "S_t = S_t-1 · exp(drift + σ · Z)",
        },
    }


def _path_window_map(payload: dict[str, Any], n_paths: int) -> dict[int, tuple[str, str]]:
    """Map path_id → (start_iso, end_iso) for Excel / preview meta columns."""
    out: dict[int, tuple[str, str]] = {}
    raw = payload.get("path_windows") or []
    if isinstance(raw, dict):
        for k, v in raw.items():
            try:
                pid = int(k)
                if isinstance(v, (list, tuple)) and len(v) >= 2:
                    out[pid] = (str(v[0])[:10], str(v[1])[:10])
                elif isinstance(v, dict):
                    out[pid] = (str(v.get("start") or "")[:10], str(v.get("end") or "")[:10])
            except Exception:
                continue
    else:
        for row in raw:
            try:
                pid = int(row["path_id"])
                out[pid] = (str(row.get("start") or "")[:10], str(row.get("end") or "")[:10])
            except Exception:
                continue
    # Fallback: blank windows so columns still exist.
    for pid in range(1, n_paths + 1):
        out.setdefault(pid, ("", ""))
    return out


def matrix_preview(
    payload: dict[str, Any],
    *,
    max_paths: int = 25,
    max_dates: int = 40,
) -> dict[str, Any]:
    """Preview without requiring the full matrix in RAM — stream GBM rows.

    When the horizon is longer than ``max_dates``, take the first half and last
    half of trading dates so Product End remains visible (not only early columns).
    """
    dates: list[date] = payload["dates"]
    mat = payload.get("matrix")
    total_paths = int(payload.get("n_paths") or (mat.shape[0] if mat is not None else 0))
    total_dates = int(payload.get("n_dates") or len(dates))
    n_paths = min(int(max_paths), total_paths)
    budget = min(int(max_dates), total_dates, len(dates))
    if total_dates <= budget:
        date_indices = list(range(total_dates))
    else:
        head = budget // 2
        tail = budget - head
        date_indices = list(range(head)) + list(range(total_dates - tail, total_dates))
    preview_dates = [dates[i] for i in date_indices]
    n_dates = len(preview_dates)
    windows = _path_window_map(payload, total_paths)
    headers = ["Path", "Start Date", "End Date"] + [d.isoformat() for d in preview_dates]
    rows: list[list[float | int | str]] = []
    base_seed = int(payload.get("base_seed") or GBM_BASE_SEED)
    if mat is not None:
        for i in range(n_paths):
            pid = i + 1
            start, end = windows.get(pid, ("", ""))
            rows.append(
                [pid, start, end]
                + [round(float(mat[i, j]), 4) for j in date_indices]
            )
    else:
        params = GbmParams(
            spot0=float(payload["spot0"]),
            asof=str(payload.get("asof") or ""),
            mean_return=float(payload["mean_return"]),
            std_dev=float(payload["std_dev"]),
            drift=float(payload["drift"]),
            n_returns=int(payload.get("n_returns") or 0),
            first_date=str(payload.get("first_date") or "2001-01-01"),
            last_date=str(payload.get("last_date") or payload.get("asof") or ""),
        )
        for path_id in range(1, n_paths + 1):
            spots = gbm_spots(
                params.spot0,
                len(dates),
                params.drift,
                params.std_dev,
                path_id=path_id,
                base_seed=base_seed,
            )
            start, end = windows.get(path_id, ("", ""))
            rows.append(
                [path_id, start, end]
                + [round(float(spots[j]), 4) for j in date_indices]
            )
    return {
        **matrix_meta(payload),
        "preview_paths": n_paths,
        "preview_dates": n_dates,
        "headers": headers,
        "rows": rows,
        "truncated": total_paths > n_paths or total_dates > n_dates,
        "date_sample": "head_tail" if total_dates > n_dates else "full",
        "horizon_start": dates[0].isoformat() if dates else None,
        "horizon_end": dates[-1].isoformat() if dates else None,
    }


def _iter_path_rows(
    *,
    matrix: np.ndarray | None,
    params: GbmParams | None,
    dates: list[date],
    n_paths: int,
    base_seed: int,
):
    """Yield path rows without requiring the full matrix in RAM."""
    n_dates = len(dates)
    if matrix is not None:
        for i in range(n_paths):
            yield i + 1, np.asarray(matrix[i, :n_dates], dtype=np.float64)
        return
    assert params is not None
    for path_id in range(1, n_paths + 1):
        yield path_id, gbm_spots(
            params.spot0,
            n_dates,
            params.drift,
            params.std_dev,
            path_id=path_id,
            base_seed=base_seed,
        )


def _xlsx_styles():
    """Shared openpyxl style objects — Primary SP desk theme."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin_gold = Side(style="thin", color=_BRAND_GOLD)
    thin_border = Side(style="thin", color=_BRAND_BORDER)
    thin_rule = Side(style="thin", color=_BRAND_RULE)

    gold_box = Border(top=thin_gold, bottom=thin_gold, left=thin_gold, right=thin_gold)
    grid_box = Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)

    return {
        "fill_gold_pale": PatternFill("solid", fgColor=_BRAND_GOLD_PALE),
        "fill_gold": PatternFill("solid", fgColor=_BRAND_GOLD),
        "fill_maroon_deep": PatternFill("solid", fgColor=_BRAND_MAROON_DEEP),
        "fill_parchment": PatternFill("solid", fgColor=_BRAND_PARCHMENT),
        "fill_ivory": PatternFill("solid", fgColor=_BRAND_IVORY),
        "fill_white": PatternFill("solid", fgColor=_BRAND_WHITE),
        "fill_footer": PatternFill("solid", fgColor=_BRAND_FOOTER),
        "fill_soft": PatternFill("solid", fgColor=_BRAND_PARCHMENT),
        "fill_alt": PatternFill("solid", fgColor=_BRAND_IVORY),
        "fill_maroon": PatternFill("solid", fgColor=_BRAND_MAROON_DEEP),
        "font_title": Font(name="Calibri", size=18, bold=True, color=_BRAND_WHITE),
        "font_sub": Font(name="Calibri", size=10, bold=True, color=_BRAND_INK),
        "font_eyebrow": Font(name="Calibri", size=8, italic=True, color=_BRAND_MUTED),
        "font_header": Font(name="Calibri", size=10, bold=True, color=_BRAND_WHITE),
        "font_label": Font(name="Calibri", size=10, bold=True, color=_BRAND_MAROON),
        "font_value": Font(name="Calibri", size=10, color=_BRAND_INK),
        "font_footer": Font(name="Calibri", size=8, italic=True, color=_BRAND_MUTED),
        "align_mid": Alignment(vertical="center", horizontal="left", indent=1, wrap_text=False),
        "align_center": Alignment(vertical="center", horizontal="center", wrap_text=True),
        "align_right": Alignment(vertical="center", horizontal="right", wrap_text=False),
        "border_gold": gold_box,
        "border_grid": grid_box,
        "border_param": gold_box,
        "border_header": gold_box,
        "border_header_outer_l": gold_box,
        "border_header_outer_r": gold_box,
        "border_footer": Border(top=thin_rule, bottom=thin_rule, left=thin_rule, right=thin_rule),
        # Legacy keys kept so callers that still reference them do not KeyError.
        "font_brand": Font(name="Calibri", size=11, bold=True, color=_BRAND_MAROON),
        "font_meta": Font(name="Calibri", size=8, italic=True, color=_BRAND_MUTED),
        "border_banner": gold_box,
        "border_gold_rule": Border(top=thin_gold, bottom=thin_gold),
    }


def _wcell(ws, value, *, font=None, fill=None, border=None, alignment=None, num_fmt=None):
    from openpyxl.cell.cell import WriteOnlyCell

    cell = WriteOnlyCell(ws, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    if alignment is not None:
        cell.alignment = alignment
    if num_fmt is not None:
        cell.number_format = num_fmt
    return cell


def _append_brand_chrome(
    ws,
    *,
    title: str,
    subtitle: str,
    meta: str,
    col_count: int,
    styles: dict,
    logo_path: Path | None,
) -> None:
    """Primary SP masthead — logo wash → gold rule → maroon title → gold subtitle → eyebrow."""
    from openpyxl.drawing.image import Image as XLImage

    n = max(4, int(col_count))
    # Cap styled chrome width so Daily grids stay memory-safe.
    banner_cols = min(n, 14)

    def filled_row(text: str, *, font, fill, height: float | None = None) -> list:
        row = []
        for c in range(1, banner_cols + 1):
            row.append(
                _wcell(
                    ws,
                    text if c == 1 else "",
                    font=font if c == 1 else styles["font_value"],
                    fill=fill,
                    alignment=styles["align_mid"],
                )
            )
        if n > banner_cols:
            row.extend([""] * (n - banner_cols))
        return row

    # Rows 1–3: gold-pale wash + ARWL logo
    for _ in range(3):
        ws.append(filled_row("", font=styles["font_value"], fill=styles["fill_gold_pale"]))
    if logo_path is not None and logo_path.is_file():
        try:
            img = XLImage(str(logo_path))
            img.width = 210
            img.height = 48
            ws.add_image(img, "A1")
        except Exception:
            pass

    # Row 4: gold accent rule
    gold = [
        _wcell(ws, "", fill=styles["fill_gold"], border=styles["border_gold_rule"])
        for _ in range(banner_cols)
    ]
    if n > banner_cols:
        gold.extend([""] * (n - banner_cols))
    ws.append(gold)

    # Row 5: maroon-deep title
    ws.append(
        filled_row(title, font=styles["font_title"], fill=styles["fill_maroon_deep"])
    )
    # Row 6: gold subtitle
    ws.append(filled_row(subtitle, font=styles["font_sub"], fill=styles["fill_gold"]))
    # Row 7: parchment eyebrow
    ws.append(
        filled_row(
            meta.strip() or _DESK_EYEBROW,
            font=styles["font_eyebrow"],
            fill=styles["fill_parchment"],
        )
    )
    # Row 8: spacer
    ws.append(filled_row("", font=styles["font_value"], fill=styles["fill_parchment"]))

    try:
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 6
        ws.row_dimensions[4].height = 5
        ws.row_dimensions[5].height = 36
        ws.row_dimensions[6].height = 22
        ws.row_dimensions[7].height = 16
        ws.row_dimensions[8].height = 8
    except Exception:
        pass


def write_mc_matrix_xlsx(
    payload: dict[str, Any],
    dest: Path,
    *,
    max_paths: int | None = None,
    params: GbmParams | None = None,
) -> Path:
    """Branded desk Excel — Primary SP masthead + Path/Start/End columns, memory-safe stream.

    Chrome (logo wash, gold rule, maroon-deep title, gold subtitle, parchment eyebrow,
    maroon headers) matches frontend/lib/download.ts. Path×date body rows are streamed
    as plain values so free-tier hosts never materialise millions of styled cells.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    dates: list[date] = list(payload["dates"])
    mat = payload.get("matrix")
    n_paths_all = int(payload.get("n_paths") or (mat.shape[0] if mat is not None else 0))
    if max_paths is not None:
        n_paths = min(n_paths_all, max(1, int(max_paths)))
    else:
        n_paths = n_paths_all
    n_dates = len(dates)
    if n_paths <= 0 or n_dates <= 0:
        raise ValueError("Empty Monte Carlo matrix — nothing to export")

    mean_ret = float(payload["mean_return"])
    std_dev = float(payload["std_dev"])
    drift = float(payload["drift"])
    spot0 = float(payload["spot0"])
    asof = str(payload.get("asof") or "")
    hist_first = str(payload.get("first_date") or "2001-01-01")
    hist_last = str(payload.get("last_date") or asof)
    base_seed = int(payload.get("base_seed") or GBM_BASE_SEED)
    cells = n_paths * n_dates
    capped_note = ""
    if n_paths < n_paths_all:
        capped_note = f"Export capped to {n_paths} of {n_paths_all} paths for memory."
    elif cells > _EXCEL_CELL_SOFT_CAP:
        capped_note = "Large grid streamed with branded desk chrome · deploy-safe."

    gbm_params = params
    if gbm_params is None and mat is None:
        gbm_params = GbmParams(
            spot0=spot0,
            asof=asof,
            mean_return=mean_ret,
            std_dev=std_dev,
            drift=drift,
            n_returns=int(payload.get("n_returns") or 0),
            first_date=hist_first,
            last_date=hist_last,
        )

    styles = _xlsx_styles()
    logo = _LOGO if _LOGO and _LOGO.is_file() else _resolve_logo()
    export_day = _desk_date(date.today())
    horizon_start = _desk_date(dates[0]) if dates else _desk_date(asof or hist_last)
    horizon_end = _desk_date(dates[-1]) if dates else ""
    subtitle_line = f"{n_paths:,} paths · {n_dates:,} trading dates · Exported {export_day}"

    # write_only = only current row in RAM — required for Daily on free Render.
    wb = Workbook(write_only=True)
    wb.creator = "Anand Rathi Wealth"
    try:
        wb.title = "Simulated Nifty Paths"
    except Exception:
        pass

    # ── Parameters ──────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Parameters")
    try:
        ws_p.sheet_properties.tabColor = _BRAND_MAROON
        ws_p.sheet_view.showGridLines = False
    except Exception:
        pass
    # Wide enough that Parameter / Value labels are never clipped.
    ws_p.column_dimensions["A"].width = 32
    ws_p.column_dimensions["B"].width = 28
    ws_p.column_dimensions["C"].width = 16
    ws_p.column_dimensions["D"].width = 14

    _append_brand_chrome(
        ws_p,
        title="Simulated Nifty Paths · Parameters",
        subtitle=subtitle_line,
        meta=capped_note or _DESK_EYEBROW,
        col_count=4,
        styles=styles,
        logo_path=logo,
    )

    header_p = [
        _wcell(
            ws_p,
            "Parameter",
            font=styles["font_header"],
            fill=styles["fill_maroon"],
            border=styles["border_header_outer_l"],
            alignment=styles["align_mid"],
        ),
        _wcell(
            ws_p,
            "Value",
            font=styles["font_header"],
            fill=styles["fill_maroon"],
            border=styles["border_header_outer_r"],
            alignment=styles["align_mid"],
        ),
        "",
        "",
    ]
    ws_p.append(header_p)

    # Desk-facing parameters only — no formula / seed / duplicate %-point rows.
    param_rows: list[tuple[str, Any, str | None]] = [
        ("Current Nifty Spot", spot0, "#,##0.00"),
        ("Daily Average Return", mean_ret, "0.0000%"),
        ("Daily Standard Deviation", std_dev, "0.00%"),
        ("Drift", drift, "0.000000"),
        ("Estimation Start", _desk_date(hist_first), None),
        ("Estimation End", _desk_date(hist_last), None),
        ("Horizon Start", horizon_start, None),
        ("Horizon End", horizon_end, None),
        ("Number Of Paths", int(n_paths), "#,##0"),
        ("Number Of Trading Dates", int(n_dates), "#,##0"),
    ]
    for i, (label, value, fmt) in enumerate(param_rows):
        alt = i % 2 == 1
        fill = styles["fill_alt"] if alt else styles["fill_white"]
        cell_val = value
        ws_p.append(
            [
                _wcell(
                    ws_p,
                    label,
                    font=styles["font_label"],
                    fill=styles["fill_soft"] if not alt else fill,
                    border=styles["border_param"],
                    alignment=styles["align_mid"],
                ),
                _wcell(
                    ws_p,
                    cell_val,
                    font=styles["font_value"],
                    fill=fill,
                    border=styles["border_param"],
                    alignment=styles["align_right"]
                    if isinstance(cell_val, (int, float))
                    else styles["align_mid"],
                    num_fmt=fmt,
                ),
                "",
                "",
            ]
        )

    ws_p.append(
        [
            _wcell(
                ws_p,
                f"Anand Rathi Wealth · Gift City AIF · {n_paths:,} paths · Exported {export_day}",
                font=styles["font_footer"],
                fill=styles["fill_footer"],
                border=styles["border_footer"],
                alignment=styles["align_mid"],
            ),
            _wcell(ws_p, "", fill=styles["fill_footer"], border=styles["border_footer"]),
            "",
            "",
        ]
    )

    # ── Simulated Nifty ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Simulated Nifty")
    try:
        ws.sheet_properties.tabColor = _BRAND_GOLD
        ws.sheet_view.showGridLines = False
    except Exception:
        pass
    # Freeze panes must be set before streaming rows on write_only sheets.
    try:
        ws.freeze_panes = f"D{_CHROME_DATA_START}"
    except Exception:
        pass
    # Keep Path / Start / End fully readable; date columns wide enough for DD-MMM-YYYY.
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    try:
        ws.sheet_format.defaultColWidth = 12
    except Exception:
        pass
    for col_idx in range(4, min(n_dates + 4, 256)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    sim_cols = n_dates + 3  # Path · Start Date · End Date · trading dates…
    windows = _path_window_map(payload, n_paths)
    _append_brand_chrome(
        ws,
        title="Simulated Nifty Paths",
        subtitle=f"{n_paths:,} paths · {n_dates:,} trading dates",
        meta=capped_note or f"Start {horizon_start} · End {horizon_end} · Exported {export_day}",
        col_count=sim_cols,
        styles=styles,
        logo_path=logo,
    )

    # Maroon header — Path · Start Date · End Date · trading dates.
    meta_headers = ["Path", "Start Date", "End Date"]
    header_labels = meta_headers + [_desk_date(d) for d in dates]
    header_cells = []
    for c, label in enumerate(header_labels, start=1):
        if c == 1:
            border = styles["border_header_outer_l"]
        elif c == sim_cols:
            border = styles["border_header_outer_r"]
        else:
            border = styles["border_header"]
        header_cells.append(
            _wcell(
                ws,
                label,
                font=styles["font_header"],
                fill=styles["fill_maroon"],
                border=border,
                alignment=styles["align_mid"],
            )
        )
    ws.append(header_cells)
    try:
        ws.row_dimensions[_CHROME_HEADER_ROW].height = 26
    except Exception:
        pass

    progress_cb = payload.get("_progress_cb")
    for idx, (path_id, spots) in enumerate(
        _iter_path_rows(
            matrix=mat if mat is not None else None,
            params=gbm_params,
            dates=dates,
            n_paths=n_paths,
            base_seed=base_seed,
        ),
        start=1,
    ):
        start_iso, end_iso = windows.get(int(path_id), ("", ""))
        start_lbl = _desk_date(start_iso) if start_iso else ""
        end_lbl = _desk_date(end_iso) if end_iso else ""
        # Plain values only — never WriteOnlyCell per grid cell (OOM on Daily).
        ws.append(
            [int(path_id), start_lbl, end_lbl] + [round(float(x), 4) for x in spots]
        )
        if progress_cb and (idx % 25 == 0 or idx == n_paths):
            try:
                progress_cb(idx / n_paths, f"Writing path {idx} of {n_paths}")
            except Exception:
                pass

    footer_cells = [
        _wcell(
            ws,
            f"Anand Rathi Wealth · Gift City AIF · {n_paths:,} paths · {n_dates:,} trading dates · Exported {export_day}",
            font=styles["font_footer"],
            fill=styles["fill_footer"],
            border=styles["border_footer"],
            alignment=styles["align_mid"],
        )
    ]
    # Pad footer visually across a few columns without styling thousands of cells.
    for _ in range(min(sim_cols - 1, 13)):
        footer_cells.append(
            _wcell(ws, "", fill=styles["fill_footer"], border=styles["border_footer"])
        )
    if sim_cols > 14:
        footer_cells.extend([""] * (sim_cols - 14))
    ws.append(footer_cells)

    try:
        ws.auto_filter.ref = f"A{_CHROME_HEADER_ROW}:{_col_letter(sim_cols)}{_CHROME_HEADER_ROW}"
    except Exception:
        pass

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def _col_letter(n: int) -> str:
    s = ""
    x = int(n)
    while x > 0:
        x, r = divmod(x - 1, 26)
        s = chr(65 + r) + s
    return s
