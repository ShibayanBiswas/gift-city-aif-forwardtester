"""Edge-case + branding checks for wind-up readiness."""
from __future__ import annotations

import sys
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from openpyxl import load_workbook

from app.engine.gbm import estimate_gbm_params
from app.engine.market import clear_market_cache, load_market
from app.engine.mc_matrix import build_mc_matrix, write_mc_matrix_xlsx
from app.engine.paths import build_forward_market
from app.engine.product import parse_product_workbook, resolved_simulation_end


def _assert(cond: bool, msg: str) -> None:
    if not cond:
        raise AssertionError(msg)


def main() -> None:
    clear_market_cache()
    m = load_market()
    gbm = estimate_gbm_params(m)

    # Weekend / holiday as-of: last_date must be a Mon–Fri session.
    _assert(m.last_date.weekday() < 5, f"asof not a weekday: {m.last_date}")
    _assert(gbm.asof == m.last_date.isoformat(), "GBM asof != market last_date")
    _assert(abs(gbm.spot0 - float(m.closes[-1])) < 1e-9, "S0 != last close")
    _assert(str(gbm.first_date).startswith("2001"), "history must start 2001")
    _assert(gbm.n_returns == len(m.closes) - 1, "μ/σ sample size mismatch")

    # Simulated "today is Sunday" → as-of stays last trading close.
    today = date.today()
    if today.weekday() >= 5:
        _assert(m.last_date < today, "weekend: asof must be prior trading day")
        print("weekend as-of OK", m.last_date, "<", today)
    else:
        # Even on a weekday, last_date is never after today.
        _assert(m.last_date <= today, "asof cannot be in the future")
        print("weekday as-of OK", m.last_date, "≤", today)

    # μ/σ dynamic: recompute after cache clear must match.
    clear_market_cache()
    gbm2 = estimate_gbm_params(load_market())
    _assert(abs(gbm2.mean_return - gbm.mean_return) < 1e-15, "μ not stable for same data")
    _assert(abs(gbm2.std_dev - gbm.std_dev) < 1e-15, "σ not stable for same data")
    print(
        "μ/σ",
        round(gbm.mean_return * 100, 4),
        "% /",
        round(gbm.std_dev * 100, 2),
        "% · S0",
        round(gbm.spot0, 2),
        "· asof",
        gbm.asof,
    )

    product = parse_product_workbook(ROOT / "Product_Input_File.xlsx")
    sim_end = resolved_simulation_end(m.last_date, product)
    fwd, params = build_forward_market(
        m,
        sim_end,
        product.tenure_days,
        observation_months=product.observation_months,
        fill_gbm=False,
    )
    _assert(params.asof == gbm.asof, "forward market params asof drift")
    # fill_gbm=False → forward dates exist but closes after asof are NaN/absent of path fill
    asof_idx = fwd.date_to_idx[m.last_date]
    _assert(abs(float(fwd.closes[asof_idx]) - gbm.spot0) < 1e-6, "desk spot at asof")
    print("forward calendar", "trading_days", len(fwd.dates), "sim_end", sim_end)

    # Branded MC Excel: logo + desk dates + soft banner sheets
    dates = [m.last_date + timedelta(days=i) for i in range(0, 14) if (m.last_date + timedelta(days=i)).weekday() < 5][:8]
    mat = build_mc_matrix(gbm, dates, 4)
    out = ROOT / "data" / "jobs" / "_edge_xlsx"
    out.mkdir(parents=True, exist_ok=True)
    payload = {
        "matrix": mat,
        "dates": dates,
        "mean_return": gbm.mean_return,
        "std_dev": gbm.std_dev,
        "drift": gbm.drift,
        "spot0": gbm.spot0,
        "asof": gbm.asof,
        "first_date": gbm.first_date,
        "last_date": gbm.last_date,
        "n_paths": 4,
        "path_windows": [
            {"path_id": 1, "start": dates[0].isoformat(), "end": dates[-1].isoformat()},
            {"path_id": 2, "start": dates[0].isoformat(), "end": dates[-1].isoformat()},
            {"path_id": 3, "start": dates[0].isoformat(), "end": dates[-1].isoformat()},
            {"path_id": 4, "start": dates[0].isoformat(), "end": dates[-1].isoformat()},
        ],
    }
    xlsx = write_mc_matrix_xlsx(payload, out / "edge_mc.xlsx")
    # data_only=False so we can inspect images / styles
    wb = load_workbook(xlsx, read_only=False, data_only=False)
    _assert("Parameters" in wb.sheetnames and "Simulated Nifty" in wb.sheetnames, wb.sheetnames)
    ws = wb["Parameters"]
    # Brand copy sits in column C under the logo (same layout as desk Excel downloads).
    brand_hit = False
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=4, values_only=True):
        if any(cell and "Anand Rathi" in str(cell) for cell in row):
            brand_hit = True
            break
    _assert(brand_hit, "brand title missing")
    _assert(len(ws._images) >= 1, "Parameters sheet missing ARWL logo")
    ws2 = wb["Simulated Nifty"]
    _assert(len(ws2._images) >= 1, "Simulated Nifty sheet missing ARWL logo")
    rows_s = list(ws2.iter_rows(min_row=1, max_row=8, values_only=True))
    header = next((r for r in rows_s if r and r[0] == "Path"), None)
    _assert(header is not None, "Path header missing")
    _assert(header[1] == "Start Date" and header[2] == "End Date", header[:4])
    _assert(
        "-" in str(header[3])
        and any(
            m in str(header[3])
            for m in ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")
        ),
        header[3],
    )
    # Maroon header fill on Path cell — desk chrome parity with download.ts
    path_header = ws2.cell(6, 1)
    fill = getattr(path_header.fill, "fgColor", None)
    rgb = getattr(fill, "rgb", None) or getattr(fill, "theme", None)
    _assert(rgb is not None and "7A1E2C" in str(rgb).upper(), f"maroon header fill missing: {rgb}")
    print("xlsx branding OK", xlsx.name, "bytes", xlsx.stat().st_size, "logos", len(ws._images), len(ws2._images))
    wb.close()

    # Independent paths: same date index → different prices
    _assert(abs(float(mat[0, 1]) - float(mat[1, 1])) > 1e-9, "paths must diverge")
    print("EDGE CASES PASS")


if __name__ == "__main__":
    main()
