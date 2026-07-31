"""Build branded Anand Rathi Product_Input_File.xlsx (percentages + Put Option labels)."""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
LOGO = ROOT / "frontend" / "public" / "brand" / "arwl-logo.png"
OUT = ROOT / "Product_Input_File.xlsx"
UPLOADS = ROOT / "data" / "uploads"

MAROON = "7A1E2C"
GOLD = "D4B24C"
HEADER_FILL = "7A1E2C"
ALT = "FAF6F0"
SOFT = "FFF8EC"
INK = "1F1612"
WHITE = "FFFFFF"

THIN = Border(
    left=Side(style="thin", color=GOLD),
    right=Side(style="thin", color=GOLD),
    top=Side(style="thin", color=GOLD),
    bottom=Side(style="thin", color=GOLD),
)

# WF1 Computation defaults — also baked into product.py / nav.py.
# kind: "int" | "pct" | "rate" | "date"
FUND_ROWS = [
    ("Principal", 1_000_000_000, "#,##0", "int"),
    ("Tenure Days", 1930, "#,##0", "int"),
    ("Cash Buffer %", 0.05, "0.00%", "pct"),
    ("G-Sec Sleeve %", 0.95, "0.00%", "pct"),
    ("Cash Interest Rate", 0.06, "0.00%", "pct"),
    ("G-Sec Interest Rate", 0.06, "0.00%", "pct"),
    ("Management Fee Rate", 0.015, "0.00%", "pct"),
    ("Buy Brokerage", 5.32155129382014e-05, "0.000000%", "rate"),
    ("Sell Brokerage", 0.000180715512938201, "0.000000%", "rate"),
    ("GST Rate", 0.18, "0.00%", "pct"),
    ("Futures Roll Rate", 0.07, "0.00%", "pct"),
    ("Tax Benefit On Roll", 0.42744, "0.00%", "pct"),
    ("Simulation End Days", 3650, "#,##0", "int"),
]


def _font(bold=False, size=11, color=INK, italic=False):
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def build() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Input"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    widths = {
        "A": 22,
        "B": 16,
        "C": 14,
        "D": 12,
        "E": 12,
        "F": 14,
        "G": 14,
        "H": 12,
        "I": 28,
        "J": 4,
        "K": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in range(1, 5):
        ws.row_dimensions[r].height = {1: 22, 2: 20, 3: 8, 4: 8}[r]

    if LOGO.exists():
        img = XLImage(str(LOGO))
        img.width = 168
        img.height = 48
        ws.add_image(img, "A1")

    ws.merge_cells("C1:I1")
    ws["C1"] = "Anand Rathi Wealth · Gift City"
    ws["C1"].font = _font(bold=True, size=12, color=MAROON)
    ws["C1"].alignment = Alignment(vertical="center")

    ws.merge_cells("C2:I2")
    ws["C2"] = "Product Input · Structured Unit Definition"
    ws["C2"].font = _font(bold=True, size=14, color=INK)

    for c in range(1, 10):
        ws.cell(4, c).fill = PatternFill("solid", fgColor=GOLD)

    # Fund economics — principal, sleeves, interest, fees, buy/sell rates.
    fund_start = 6
    for i, (label, value, fmt, kind) in enumerate(FUND_ROWS):
        r = fund_start + i
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = _font(bold=True, color=MAROON)
        ws.cell(r, 1).fill = PatternFill("solid", fgColor=SOFT)
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2).value = value
        ws.cell(r, 2).font = _font(bold=True)
        ws.cell(r, 2).border = THIN
        ws.cell(r, 2).alignment = Alignment(horizontal="center")
        if kind == "date":
            ws.cell(r, 2).number_format = "YYYY-MM-DD"
        else:
            ws.cell(r, 2).number_format = fmt

    # Observation months only — calendar days (m × 30.5) are computed in the web app.
    ws["K6"] = "Observation Months"
    ws["K6"].font = _font(bold=True, size=10, color=WHITE)
    ws["K6"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["K6"].alignment = Alignment(horizontal="center")
    ws["K6"].border = THIN
    obs = [38, 41, 44, 47, 50, 53, 56]
    for i, m in enumerate(obs):
        r = 7 + i
        ws.cell(r, 11, m).border = THIN
        ws.cell(r, 11).alignment = Alignment(horizontal="center")
        ws.cell(r, 11).font = _font()
        if i % 2:
            ws.cell(r, 11).fill = PatternFill("solid", fgColor=ALT)

    headers = [
        "Return Level",
        "Strike %",
        "Option",
        "Forward",
        "Discount",
        "Vol Near",
        "Vol Far",
        "Qty",
        "Include",
    ]
    header_row = fund_start + len(FUND_ROWS) + 1  # blank row after fund block
    ws.row_dimensions[header_row].height = 28
    for c, h in enumerate(headers, 1):
        cell = ws.cell(header_row, c, h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

    legs = [
        (0.37, 1.37, "Put Option", 0.066, 0.076, 0.14368857564668683, 0.14793275750723667, -91.5, "Yes"),
        (0.36, 1.36, "Put Option", 0.066, 0.076, 0.1441638312604284, 0.14848846191959258, 90.5, "Yes"),
        (0.25, 1.25, "Put Option", 0.066, 0.076, 0.15118881279350588, 0.15619717745743442, 1.0, "Yes"),
        (-0.15, 0.85, "Put Option", 0.066, 0.076, 0.20241167320096648, 0.2061305717267251, -25.6, "Yes"),
        (-0.16, 0.84, "Put Option", 0.066, 0.076, 0.2041294400033543, 0.20774616738378116, 24.0, "Yes"),
        (-0.30, 0.70, "Put Option", 0.066, 0.076, 0.23005561163223529, 0.2319842472882776, 1.0, "Yes"),
    ]

    pct_cols = {1, 2, 4, 5, 6, 7}
    for i, leg in enumerate(legs):
        r = header_row + 1 + i
        ws.row_dimensions[r].height = 20
        for c, v in enumerate(leg, 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = _font(bold=(c == 8))
            if i % 2:
                cell.fill = PatternFill("solid", fgColor=ALT)
            if c in pct_cols and isinstance(v, float):
                cell.number_format = "0.00%"
            if c == 8 and isinstance(v, (int, float)):
                cell.number_format = "0.0"

    wb.save(OUT)
    UPLOADS.mkdir(parents=True, exist_ok=True)
    for name in ("current_product.xlsx", "default_product.xlsx"):
        shutil.copy2(OUT, UPLOADS / name)
    # Keep a copy under backend/ so Render (root directory = backend) always ships the sample.
    backend_copy = ROOT / "backend" / "Product_Input_File.xlsx"
    shutil.copy2(OUT, backend_copy)
    return OUT


if __name__ == "__main__":
    path = build()
    print("Wrote", path, "size", path.stat().st_size)
