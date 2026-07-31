#!/usr/bin/env python3
"""Sync market calendars through present.

1. Optional: seed roll/expiry overrides from Gift AIF Working File 1.xlsm
2. Append Nifty daily closes via Yahoo (^NSEI)
3. Extend futures roll shifts + monthly expiries through last Nifty date

Usage:
  PYTHONPATH=backend python3 scripts/sync_market_data.py
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

from app.engine.calendar_build import write_expiries_csv  # noqa: E402
from app.engine.market import DATA, _to_date, clear_market_cache, market_meta  # noqa: E402
from app.engine.market_sync import sync_market_to_present  # noqa: E402


def extract_excel_calendars(xlsm: Path) -> tuple[list, list[tuple]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsm, data_only=True, read_only=True)
    opt = []
    ws = wb["Expiry"]
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[1] is None:
            break
        opt.append(_to_date(row[1]))

    rolls = []
    ws = wb["Roll Cost + Paths"]
    for row in ws.iter_rows(min_row=3, max_col=3, values_only=True):
        if row[1] is None:
            break
        rolls.append((_to_date(row[1]), float(row[2])))
    wb.close()
    return opt, rolls


def main() -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    xlsm_candidates = [
        ROOT / "Gift AIF Working File 1.xlsm",
        Path(r"C:\Users\shiba\OneDrive\Desktop\Files\Gift AIF Working File 1.xlsm"),
    ]
    xlsm = next((p for p in xlsm_candidates if p.exists()), None)
    if xlsm is not None:
        opt, rolls = extract_excel_calendars(xlsm)
        overrides = DATA / "expiry_overrides.csv"
        write_expiries_csv(overrides, opt)
        print(f"Wrote {overrides} ({len(opt)} Excel option expiries) from {xlsm.name}")

        roll_path = DATA / "roll_costs.csv"
        with roll_path.open("w") as f:
            f.write("expiry_date,roll_cost\n")
            for d, c in rolls:
                f.write(f"{d.isoformat()},{c}\n")
        print(f"Seeded {roll_path} ({len(rolls)} Excel futures shifts — will extend to present)")
    else:
        print("Working File 1.xlsm not found — extending existing CSVs to present.")

    clear_market_cache()
    result = sync_market_to_present(force=True)
    meta = result.get("market") or market_meta()
    print(
        f"Nifty {meta.get('first_date')} → {meta.get('last_date')} · "
        f"{meta.get('trading_days')} days · "
        f"expiries={meta.get('expiries')} ({meta.get('last_expiry')}) · "
        f"rolls={meta.get('roll_shifts')} ({meta.get('last_roll_shift')})"
    )
    print(f"Yahoo rows added: {result.get('nifty', {}).get('added', 0)}")


if __name__ == "__main__":
    main()
