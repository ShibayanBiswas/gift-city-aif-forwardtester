"""Parse Product Input Excel into a structured product definition."""
from __future__ import annotations

import os
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

# Working File 1 As-per-HS defaults (Path-invariant).
DEFAULT_FORWARD = 0.066
DEFAULT_DISCOUNT = 0.076

# Fund economics — WF1 Computation defaults (desk-editable via Product Input).
DEFAULT_CASH_PCT = 0.05
DEFAULT_GSEC_PCT = 0.95
DEFAULT_CASH_RATE = 0.06
DEFAULT_GSEC_RATE = 0.06
DEFAULT_FEE_RATE = 0.015
DEFAULT_BUY_BROKERAGE = 5.32155129382014e-05
DEFAULT_SELL_BROKERAGE = 0.000180715512938201
# Aliases — ProductSpec keeps these equal to brokerage (single Tx card throughout).
DEFAULT_BUY_RATE = DEFAULT_BUY_BROKERAGE
DEFAULT_SELL_RATE = DEFAULT_SELL_BROKERAGE
DEFAULT_ROLL_RATE = 0.07
DEFAULT_TAX_BENEFIT_RATE = 0.42744
# WF1 Computation AG = AF × 18% (cash-side GST). AF brokerage is currently always 0.
DEFAULT_CASH_GST_RATE = 0.18
# Legacy only — nav ignores switch; Tx always uses buy/sell brokerage.
DEFAULT_RATE_SWITCH_DATE = date(2024, 10, 31)
# roll_costs.csv is generated at this assumed futures carry rate.
ROLL_COST_BASE_RATE = 0.07

# Forward-tester product window ends at tenure from as-of (no separate Simulation End Days).
# Monte Carlo path count over that single window (path_id 1…N, independent GBM seeds).
DEFAULT_N_PATHS = 1000
MIN_N_PATHS = 1
MAX_N_PATHS = 10000


def _add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        # 29-Feb → 28-Feb on non-leap target years
        return d.replace(month=2, day=28, year=d.year + years)


def path_end_calendar(start: date, tenure_days: int | None = None) -> date:
    """Excel / backtester tenure end rule (verbatim) — product forward window end."""
    if tenure_days is not None and not (1700 <= int(tenure_days) <= 2000):
        return start + timedelta(days=int(tenure_days))
    anniversary = _add_years(start, 5)
    return anniversary.replace(day=1) - timedelta(days=1)


def resolved_n_paths(product: "ProductSpec | None" = None, explicit: int | None = None) -> int:
    """Monte Carlo path count for the single as-of → tenure window."""
    if explicit is not None:
        n = int(explicit)
    elif product is not None and getattr(product, "n_paths", None) is not None:
        n = int(product.n_paths)  # type: ignore[arg-type]
    else:
        raw = os.environ.get("FORWARDTEST_N_PATHS") or os.environ.get("N_PATHS")
        n = int(raw) if raw and str(raw).strip().isdigit() else DEFAULT_N_PATHS
    return max(MIN_N_PATHS, min(MAX_N_PATHS, n))


def resolved_simulation_end_days(product: "ProductSpec | None" = None) -> int:
    """Calendar span of the product forward window (tenure days). Kept for API compat."""
    if product is not None and product.tenure_days > 0:
        return int(product.tenure_days)
    return 1930


def resolved_simulation_end(asof: date, product: "ProductSpec | None" = None) -> date:
    """Product End = as-of + tenure (Backtester anniversary rule when tenure ∈ [1700,2000])."""
    tenure = int(product.tenure_days) if product is not None and product.tenure_days > 0 else 1930
    return path_end_calendar(asof, tenure)


# Desk product catalogue: observation *count* is 1…7 (month offsets still float).
MIN_OBSERVATION_COUNT = 1
MAX_OBSERVATION_COUNT = 7
# Month offset band (×30.5 days from path start).
_OBS_MONTH_MIN = 1e-6
_OBS_MONTH_MAX = 120.0


def normalize_observation_months(months: list[float] | None) -> list[float]:
    """Deduplicate (order-preserving) and validate observation month offsets.

    Supported product variety: **1 to 7** observation months (not fixed at 7).
    Path count, hedge leg expansion, and contract sizing all key off this list.
    """
    if not months:
        raise ValueError(
            "Product must define at least one observation month "
            "(Product Input column whose header contains 'observation')"
        )
    out: list[float] = []
    seen: set[float] = set()
    for raw in months:
        m = float(raw)
        if not (_OBS_MONTH_MIN < m <= _OBS_MONTH_MAX):
            raise ValueError(
                f"Observation month {m} out of range ({_OBS_MONTH_MIN}, {_OBS_MONTH_MAX}]"
            )
        if m in seen:
            continue
        seen.add(m)
        out.append(m)
    if not out:
        raise ValueError("Product must define at least one observation month")
    n = len(out)
    if n < MIN_OBSERVATION_COUNT or n > MAX_OBSERVATION_COUNT:
        raise ValueError(
            f"Observation count must be between {MIN_OBSERVATION_COUNT} and "
            f"{MAX_OBSERVATION_COUNT} (got {n}). Supported desk products use 1…7 observations."
        )
    return out


@dataclass
class OptionLegSpec:
    return_level: float
    strike_pct: float
    quantity: float
    option_type: str = "P"  # Put
    forward_rate: float = DEFAULT_FORWARD
    discount_rate: float = DEFAULT_DISCOUNT
    vol: float = 0.15  # far-observation vol (HS col H for obs index ≥ 1)
    vol_near: float | None = None  # first observation vol (HS col H for obs index 0)
    include: bool = True  # False = full-hedge / display-only (e.g. qty 0.6)

    def vol_for_observation(self, obs_index: int) -> float:
        if obs_index == 0 and self.vol_near is not None:
            return float(self.vol_near)
        return float(self.vol)


@dataclass
class ProductSpec:
    name: str
    principal: float
    tenure_days: int
    observation_months: list[float]
    legs: list[OptionLegSpec]
    source_file: str = ""
    # Fund economics (fractions). Cash / G-Sec sleeves scale with principal.
    cash_pct: float = DEFAULT_CASH_PCT
    gsec_pct: float = DEFAULT_GSEC_PCT
    cash_rate: float = DEFAULT_CASH_RATE
    gsec_rate: float = DEFAULT_GSEC_RATE
    fee_rate: float = DEFAULT_FEE_RATE
    buy_rate: float = DEFAULT_BUY_RATE
    buy_brokerage: float = DEFAULT_BUY_BROKERAGE
    sell_rate: float = DEFAULT_SELL_RATE
    sell_brokerage: float = DEFAULT_SELL_BROKERAGE
    roll_rate: float = DEFAULT_ROLL_RATE
    tax_benefit_rate: float = DEFAULT_TAX_BENEFIT_RATE
    # Cash-column GST (WF1 AG = AF × cash_gst_rate).
    cash_gst_rate: float = DEFAULT_CASH_GST_RATE
    # Legacy — unused by nav (brokerage card throughout).
    rate_switch_date: date = DEFAULT_RATE_SWITCH_DATE
    # Legacy Product Input field — ignored for horizon (window = as-of → tenure end).
    simulation_end_days: int | None = None
    # Optional Monte Carlo path count override (else FORWARDTEST_N_PATHS / default 1000).
    n_paths: int | None = None

    def __post_init__(self) -> None:
        self.observation_months = normalize_observation_months(self.observation_months)
        if self.tenure_days <= 0:
            raise ValueError(f"tenure_days must be positive, got {self.tenure_days}")
        if self.principal <= 0:
            raise ValueError(f"principal must be positive, got {self.principal}")
        if not self.active_legs:
            raise ValueError("Product must have at least one active option leg")
        self.cash_pct = float(self.cash_pct)
        self.gsec_pct = float(self.gsec_pct)
        if not (0.0 < self.cash_pct < 1.0):
            raise ValueError(f"cash_pct must be in (0, 1), got {self.cash_pct}")
        if not (0.0 < self.gsec_pct < 1.0):
            raise ValueError(f"gsec_pct must be in (0, 1), got {self.gsec_pct}")
        # Prefer cash sleeve; G-Sec is the residual so 5/95 always tracks principal.
        if abs(self.cash_pct + self.gsec_pct - 1.0) > 1e-9:
            self.gsec_pct = 1.0 - self.cash_pct
        for name, val in (
            ("cash_rate", self.cash_rate),
            ("gsec_rate", self.gsec_rate),
            ("fee_rate", self.fee_rate),
            ("buy_rate", self.buy_rate),
            ("buy_brokerage", self.buy_brokerage),
            ("sell_rate", self.sell_rate),
            ("sell_brokerage", self.sell_brokerage),
            ("roll_rate", self.roll_rate),
            ("tax_benefit_rate", self.tax_benefit_rate),
            ("cash_gst_rate", self.cash_gst_rate),
        ):
            if float(val) < 0:
                raise ValueError(f"{name} must be non-negative, got {val}")
        # Single Tx card: buy_rate/sell_rate aliases stay aligned to brokerage.
        self.buy_rate = float(self.buy_brokerage)
        self.sell_rate = float(self.sell_brokerage)
        if isinstance(self.rate_switch_date, datetime):
            self.rate_switch_date = self.rate_switch_date.date()
        elif isinstance(self.rate_switch_date, str):
            self.rate_switch_date = date.fromisoformat(self.rate_switch_date[:10])
        elif not isinstance(self.rate_switch_date, date):
            raise ValueError(f"rate_switch_date must be a date, got {self.rate_switch_date!r}")
        if self.n_paths is not None:
            self.n_paths = resolved_n_paths(explicit=int(self.n_paths))
        # Keep parsed Simulation End Days if present, but never let it drive the horizon.
        if self.simulation_end_days is not None:
            self.simulation_end_days = int(self.simulation_end_days)
            if self.simulation_end_days <= 0:
                self.simulation_end_days = None

    @property
    def principal_cr(self) -> float:
        return self.principal / 1e7

    @property
    def cash_buffer_cr(self) -> float:
        return self.principal_cr * self.cash_pct

    @property
    def gsec_opening_cr(self) -> float:
        return self.principal_cr * self.gsec_pct

    @property
    def buy_gst(self) -> float:
        """Legacy residual (buy_rate − buy_brokerage). Zero when aliases are aligned."""
        return float(self.buy_rate - self.buy_brokerage)

    @property
    def sell_gst(self) -> float:
        """Legacy residual (sell_rate − sell_brokerage). Zero when aliases are aligned."""
        return float(self.sell_rate - self.sell_brokerage)

    @property
    def n_obs(self) -> int:
        return len(self.observation_months)

    @property
    def last_observation_month(self) -> float:
        return max(self.observation_months)

    @property
    def active_legs(self) -> list[OptionLegSpec]:
        return [lg for lg in self.legs if lg.include and lg.quantity != 0]

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["rate_switch_date"] = self.rate_switch_date.isoformat()
        # Product window = tenure days from as-of (Simulation End Days no longer drives horizon).
        d["simulation_end_days"] = resolved_simulation_end_days(self)
        d["simulation_end_days_source"] = "tenure"
        d["n_paths"] = resolved_n_paths(self)
        d["principal_cr"] = self.principal_cr
        d["cash_buffer_cr"] = self.cash_buffer_cr
        d["gsec_opening_cr"] = self.gsec_opening_cr
        d["buy_gst"] = self.buy_gst
        d["sell_gst"] = self.sell_gst
        d["n_obs"] = self.n_obs
        d["last_observation_month"] = self.last_observation_month
        return d

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> ProductSpec:
        legs_raw = d.get("legs") or []
        legs: list[OptionLegSpec] = []
        for lg in legs_raw:
            legs.append(
                OptionLegSpec(
                    return_level=float(lg["return_level"]),
                    strike_pct=float(lg["strike_pct"]),
                    quantity=float(lg["quantity"]),
                    option_type=str(lg.get("option_type") or "P"),
                    forward_rate=float(lg.get("forward_rate", DEFAULT_FORWARD)),
                    discount_rate=float(lg.get("discount_rate", DEFAULT_DISCOUNT)),
                    vol=float(lg.get("vol", 0.15)),
                    vol_near=(
                        float(lg["vol_near"])
                        if lg.get("vol_near") is not None
                        else None
                    ),
                    include=bool(lg.get("include", True)),
                )
            )
        principal = d.get("principal")
        if principal is None:
            principal = float(d.get("principal_cr", 100.0)) * 1e7
        kwargs: dict[str, Any] = {
            "name": str(d.get("name") or "Current Product"),
            "principal": float(principal),
            "tenure_days": int(d["tenure_days"]),
            "observation_months": [float(x) for x in (d.get("observation_months") or [])],
            "legs": legs,
            "source_file": str(d.get("source_file") or ""),
        }
        for key in (
            "cash_pct",
            "gsec_pct",
            "cash_rate",
            "gsec_rate",
            "fee_rate",
            "buy_rate",
            "buy_brokerage",
            "sell_rate",
            "sell_brokerage",
            "roll_rate",
            "tax_benefit_rate",
            "cash_gst_rate",
        ):
            if key in d and d[key] is not None:
                kwargs[key] = float(d[key])
        if d.get("rate_switch_date") is not None:
            raw = d["rate_switch_date"]
            if isinstance(raw, datetime):
                kwargs["rate_switch_date"] = raw.date()
            elif isinstance(raw, date):
                kwargs["rate_switch_date"] = raw
            else:
                kwargs["rate_switch_date"] = date.fromisoformat(str(raw)[:10])
        if d.get("simulation_end_days") is not None:
            kwargs["simulation_end_days"] = int(d["simulation_end_days"])
        elif d.get("simulation_end_date") is not None:
            # Legacy workbook/API date → approximate days from a synthetic as-of at resolve time
            # is handled by callers; if an int-like string sneaks in, parse as days.
            raw = d["simulation_end_date"]
            try:
                kwargs["simulation_end_days"] = int(float(str(raw)))
            except (TypeError, ValueError):
                pass
        return cls(**kwargs)


def _cell_str(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _to_float(v: Any) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("%", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_rate_fraction(v: Any) -> float | None:
    """
    Parse a rate that may be Excel % (0.066), percent points (6.6), or '6.6%'.
    Returns a fraction (0.066 for 6.6%).
    """
    if v is None or isinstance(v, bool):
        return None
    has_pct = isinstance(v, str) and "%" in v
    n = _to_float(v)
    if n is None:
        return None
    if has_pct or abs(n) > 1.5:
        return n / 100.0
    return n


def _to_strike_pct(v: Any, return_level: float | None = None) -> float:
    """
    Parse strike as percent-of-spot points (137 for 137%).
    Accepts Excel % (1.37), points (137), or '137%'.
    """
    if v is None or isinstance(v, bool):
        if return_level is not None:
            return (1.0 + return_level) * 100.0
        raise ValueError("strike missing")
    has_pct = isinstance(v, str) and "%" in v
    n = _to_float(v)
    if n is None:
        if return_level is not None:
            return (1.0 + return_level) * 100.0
        raise ValueError("strike missing")
    # Excel percent cell for 137% → 1.37; bare 137 stays 137.
    if has_pct:
        return n if abs(n) > 3 else n * 100.0
    if 0 < abs(n) <= 3:
        return n * 100.0
    return n


def _option_code(v: Any) -> str:
    raw = _cell_str(v).upper()
    if not raw:
        return "P"
    if raw.startswith("C") or "CALL" in raw:
        return "C"
    return "P"


def _match_fund_label(label: str) -> str | None:
    """Map a Product Input row label to a ProductSpec economics field."""
    s = label.lower().strip()
    if not s:
        return None
    if "principal" in s:
        return "principal"
    if "tenure" in s:
        return "tenure"
    # Ignore absolute Day-0 crore rows — sleeves come from Cash Buffer % / G-Sec Sleeve %.
    if ("day 0" in s or "day0" in s or "opening" in s) and (
        "cash" in s or "g-sec" in s or "gsec" in s
    ):
        return None
    # GST before generic cash-rate (otherwise "Cash GST Rate" → cash_rate).
    if "cash gst" in s or s.strip() in {"gst", "gst rate", "gst %"} or (
        "gst" in s and "rate" in s
    ):
        return "cash_gst_rate"
    if (
        "rate switch" in s
        or "brokerage from" in s
        or "brokerage after" in s
        or ("switch" in s and "date" in s)
        or s.strip() in {"ak2"}
    ):
        return "rate_switch_date"
    # Monte Carlo path count (preferred over legacy Simulation End Days).
    if (
        "monte carlo" in s
        or "mc paths" in s
        or s.strip() in {"n paths", "n_paths", "paths", "path count", "number of paths"}
        or ("path" in s and ("count" in s or "number" in s or "n " in s or s.strip().startswith("n ")))
    ):
        return "n_paths"
    # Legacy label — parsed for workbook compat but ignored as horizon (tenure end wins).
    if (
        "simulation end days" in s
        or "horizon days" in s
        or "forward days" in s
        or ("simulation" in s and "end" in s and "day" in s)
        or ("horizon" in s and "day" in s)
        or s.strip() in {"simulation end", "horizon end", "forward end"}
    ):
        return "simulation_end_days"
    if "cash" in s and ("buffer" in s or "sleeve" in s or "%" in s or "percent" in s or "pct" in s):
        if "interest" in s or ("rate" in s and "gst" not in s):
            return "cash_rate"
        return "cash_pct"
    if "cash" in s and ("interest" in s or "rate" in s) and "gst" not in s:
        return "cash_rate"
    if ("g-sec" in s or "gsec" in s or "gov" in s) and (
        "sleeve" in s or "%" in s or "percent" in s or "pct" in s or "allocation" in s
    ):
        return "gsec_pct"
    if ("g-sec" in s or "gsec" in s or "gov" in s) and ("interest" in s or "rate" in s):
        return "gsec_rate"
    if "management" in s and "fee" in s:
        return "fee_rate"
    if s.startswith("fee") or "fee rate" in s:
        return "fee_rate"
    if "buy" in s and "broker" in s:
        return "buy_brokerage"
    if "sell" in s and "broker" in s:
        return "sell_brokerage"
    if "buy" in s and "rate" in s:
        return "buy_rate"
    if "sell" in s and "rate" in s:
        return "sell_rate"
    if "roll" in s and "rate" in s:
        return "roll_rate"
    if "futures roll" in s or (s.startswith("roll") and "cost" not in s):
        return "roll_rate"
    if "tax benefit" in s or "tax rate" in s:
        return "tax_benefit_rate"
    return None


def _to_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s[:10])
        except ValueError:
            return None
    return None


def parse_product_workbook(path: str | Path, name: str | None = None) -> ProductSpec:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = None
    for s in wb.sheetnames:
        if "product" in s.lower() or "input" in s.lower() or "as per" in s.lower():
            sheet_name = s
            break
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    grid: list[list[Any]] = []
    for row in ws.iter_rows(max_row=120, max_col=14, values_only=True):
        grid.append(list(row))
    wb.close()

    principal = None
    tenure = None
    observation_months: list[float] = []
    legs: list[OptionLegSpec] = []
    economics: dict[str, Any] = {}

    for row in grid:
        a, b = (row[0] if len(row) > 0 else None), (row[1] if len(row) > 1 else None)
        label = _cell_str(a)
        key = _match_fund_label(label)
        if key is None:
            continue
        if key == "principal":
            pv = _to_float(b)
            if pv is not None:
                principal = float(pv)
            continue
        if key == "tenure":
            tv = _to_float(b)
            if tv is not None:
                tenure = int(tv)
            continue
        if key == "rate_switch_date":
            parsed_date = _to_date(b)
            if parsed_date is not None:
                economics[key] = parsed_date
            continue
        if key == "n_paths":
            if isinstance(b, (datetime, date)) and not isinstance(b, bool):
                continue
            parsed_n = _to_float(b)
            if parsed_n is not None and parsed_n > 0:
                economics["n_paths"] = resolved_n_paths(explicit=int(round(parsed_n)))
            continue
        if key == "simulation_end_days":
            # Legacy only — horizon is tenure via path_end_calendar; keep value for display.
            if isinstance(b, (datetime, date)) and not isinstance(b, bool):
                continue
            parsed_days = _to_float(b)
            if parsed_days is not None and parsed_days > 0:
                economics[key] = int(round(parsed_days))
            continue
        # Allocation / interest / fee / tx rates — accept % cells or fractions.
        if key in {
            "cash_pct",
            "gsec_pct",
            "cash_rate",
            "gsec_rate",
            "fee_rate",
            "roll_rate",
            "tax_benefit_rate",
            "cash_gst_rate",
        }:
            parsed = _to_rate_fraction(b)
            if parsed is not None:
                economics[key] = parsed
            continue
        # Buy/sell all-in and brokerage are tiny fractions (e.g. 5.5e-5) — never ÷100.
        parsed = _to_float(b)
        if parsed is not None:
            # If desk typed 0.005499 as percent points of notional, convert.
            if key in {"buy_rate", "sell_rate", "buy_brokerage", "sell_brokerage"} and abs(parsed) > 0.01:
                parsed = parsed / 100.0
            economics[key] = float(parsed)

    # Observation months: prefer column with "observation" header
    obs_col = 7
    for row in grid[:30]:
        for c, val in enumerate(row):
            if val is not None and "observation" in str(val).lower():
                obs_col = c
                break

    for row in grid:
        if len(row) <= obs_col:
            continue
        v = row[obs_col]
        if isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= float(v) <= 120:
            observation_months.append(float(v))

    # Options book table: detect header row with QTY / Quantity
    header_row = None
    col_map: dict[str, int] = {}
    for r, row in enumerate(grid[:40]):
        labels = {_cell_str(v).lower(): c for c, v in enumerate(row) if v is not None}
        if "qty" in labels or "quantity" in labels:
            header_row = r
            col_map = labels
            break

    def col(*names: str, default: int | None = None) -> int | None:
        for n in names:
            if n in col_map:
                return col_map[n]
        return default

    qty_c = col("qty", "quantity", default=4)
    ret_c = col(
        "nifty returns",
        "return level",
        "return",
        "underlying return level",
        default=0,
    )
    strike_c = col("strike %", "strike pct", "strike as percent of spot", "strike%")
    opt_c = col("option", "option type", "type")
    fwd_c = col("forward", "forward rate")
    disc_c = col("discount", "discount rate")
    vol_c = col("vol", "vols", "volatility", "vol (far)", "vol far")
    vol_near_c = col("vol near", "vol (near)", "vol near obs")
    include_c = col("include", "active", "use")

    if header_row is not None and qty_c is not None and ret_c is not None:
        for row in grid[header_row + 1 :]:
            if len(row) <= max(qty_c, ret_c):
                continue
            qty = _to_float(row[qty_c])
            ret = _to_rate_fraction(row[ret_c])
            if qty is None or ret is None:
                continue
            # Skip blank / note rows and non-leg return noise
            if ret < -1.5 or ret > 3:
                continue

            strike_raw = row[strike_c] if strike_c is not None else None
            try:
                strike_pct = _to_strike_pct(strike_raw, return_level=ret)
            except ValueError:
                continue

            opt = "P"
            if opt_c is not None:
                opt = _option_code(row[opt_c])

            fwd = DEFAULT_FORWARD
            if fwd_c is not None and row[fwd_c] is not None:
                parsed = _to_rate_fraction(row[fwd_c])
                if parsed is not None:
                    fwd = parsed

            disc = DEFAULT_DISCOUNT
            if disc_c is not None and row[disc_c] is not None:
                parsed = _to_rate_fraction(row[disc_c])
                if parsed is not None:
                    disc = parsed

            vol = 0.15
            if vol_c is not None and row[vol_c] is not None:
                parsed = _to_rate_fraction(row[vol_c])
                if parsed is not None:
                    vol = parsed

            vol_near = None
            if vol_near_c is not None and row[vol_near_c] is not None:
                parsed = _to_rate_fraction(row[vol_near_c])
                if parsed is not None:
                    vol_near = parsed

            include = True
            if include_c is not None and row[include_c] is not None:
                raw = _cell_str(row[include_c]).lower()
                include = not (
                    raw in {"no", "n", "0", "false", "exclude", "excluded"}
                    or raw.startswith("no")
                    or "exclude" in raw
                )
            elif abs(float(qty) - 0.6) < 1e-9:
                # Legacy sheets: qty 0.6 full-hedge is display-only unless Include is set.
                include = False

            legs.append(
                OptionLegSpec(
                    return_level=float(ret),
                    strike_pct=float(strike_pct),
                    quantity=float(qty),
                    option_type=opt,
                    forward_rate=fwd,
                    discount_rate=disc,
                    vol=vol,
                    vol_near=vol_near,
                    include=include,
                )
            )
    else:
        # Legacy layout: return in A, qty in E
        for row in grid:
            if len(row) <= 4:
                continue
            ret = _to_rate_fraction(row[0])
            qty = _to_float(row[4])
            if ret is None or qty is None:
                continue
            if ret < -1.5 or ret > 3:
                continue
            include = abs(float(qty) - 0.6) >= 1e-9
            legs.append(
                OptionLegSpec(
                    return_level=float(ret),
                    strike_pct=(1.0 + float(ret)) * 100.0,
                    quantity=float(qty),
                    include=include,
                )
            )

    if principal is None:
        raise ValueError("Principal not found in product input")
    if tenure is None:
        raise ValueError("Tenure Days not found in product input")
    if not observation_months:
        raise ValueError("No observation months found")
    if not any(lg.include and lg.quantity != 0 for lg in legs):
        raise ValueError("No active option quantities found")

    return ProductSpec(
        name=name or path.stem,
        principal=principal,
        tenure_days=tenure,
        observation_months=normalize_observation_months(observation_months),
        legs=legs,
        source_file=str(path),
        **economics,
    )
